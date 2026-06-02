#!/usr/bin/env python3
"""
YES24 베스트셀러 → 저자 목록 JSON 빌드
입력: data/yes24/archive.json (generate_report.py가 매일 수집)
      또는 로컬 xlsx 파일 (폴백)
출력: data/authors/authors.json, authors.js, panels/panel24/authors-data.js
"""
import json
import os
import re
import sys

# Windows cp949 인코딩 에러 방지
if sys.stdout and sys.stdout.encoding and sys.stdout.encoding.lower().startswith("cp"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ARCHIVE_PATH = os.path.join(SCRIPT_DIR, "..", "data", "yes24", "archive.json")
OUT_DIR = os.path.join(SCRIPT_DIR, "..", "data", "authors")
PANEL_JS_PATH = os.path.join(SCRIPT_DIR, "..", "panels", "panel24", "authors-data.js")


def parse_author_field(raw: str) -> list[dict]:
    """저자 필드 파싱 → [{name, role}] 반환
    예: '아르빈드 나라야난,사야시 카푸르 저/강미경 역' → 저자 2명 + 역자 1명
    """
    if not raw:
        return []
    results = []
    # '/' 로 역할 구분 (저/역/감수 등)
    parts = raw.split('/')
    for part in parts:
        part = part.strip()
        # 역할 추출
        role = '저'
        for r in ['공저', '저', '역', '감수', '그림', '편', '원작', '사진', '엮음', '해설']:
            if part.endswith(' ' + r) or part.endswith(r):
                role = r if r != '공저' else '저'
                part = part[:part.rfind(r)].strip()
                break
        # 쉼표로 복수 저자 분리
        names = [n.strip() for n in part.split(',')]
        for name in names:
            name = name.strip()
            if not name or len(name) < 2:
                continue
            results.append({'name': name, 'role': role})
    return results


def build_from_archive(archive: dict) -> dict:
    """archive.json의 snapshots를 읽어 저자별 데이터를 집계한다."""
    author_map = {}  # name → {books: {key→{title,pub,bestRank,dates,appearances}}}
    row_count = 0

    for date_str, items in archive.get("snapshots", {}).items():
        for item in items:
            row_count += 1
            raw_author = item.get("author", "").strip()
            title = item.get("title", "").strip()
            pub = item.get("publisher", "").strip()
            rank = item.get("rank", 0)
            if isinstance(rank, str):
                try:
                    rank = int(float(rank))
                except (TypeError, ValueError):
                    rank = 0

            parsed = parse_author_field(raw_author)
            for p in parsed:
                if p['role'] in ('역', '감수', '그림', '사진', '편', '원작', '엮음', '해설'):
                    continue  # 역자/감수자 등은 제외

                name = p['name']
                book_key = title  # archive에 ISBN이 없으므로 제목 기준

                if name not in author_map:
                    author_map[name] = {'books': {}}

                if book_key not in author_map[name]['books']:
                    author_map[name]['books'][book_key] = {
                        'title': title,
                        'pub': pub,
                        'bestRank': rank if rank > 0 else 999,
                        'dates': [],
                        'appearances': 0
                    }

                bk = author_map[name]['books'][book_key]
                if rank > 0 and rank < bk['bestRank']:
                    bk['bestRank'] = rank
                bk['appearances'] += 1
                if date_str and date_str not in bk['dates']:
                    bk['dates'].append(date_str)

    # 결과 정리
    authors = []
    for name, data in author_map.items():
        books = []
        for bk_key, bk in data['books'].items():
            books.append({
                'title': bk['title'],
                'pub': bk['pub'],
                'bestRank': bk['bestRank'],
                'days': bk['appearances'],
                'lastDate': sorted(bk['dates'])[-1] if bk['dates'] else ''
            })
        books.sort(key=lambda b: b['bestRank'])
        pubs = sorted(set(b['pub'] for b in books if b['pub']))
        best_rank = min(b['bestRank'] for b in books) if books else 999
        total_days = sum(b['days'] for b in books)

        authors.append({
            'name': name,
            'books': books,
            'pubs': pubs,
            'count': len(books),
            'bestRank': best_rank,
            'totalDays': total_days,
        })

    # 권수 → 총등장일 → 최고순위 순 정렬
    authors.sort(key=lambda a: (-a['count'], -a['totalDays'], a['bestRank']))

    return {
        'authors': authors,
        'totalRows': row_count,
        'totalAuthors': len(authors),
        'totalBooks': sum(a['count'] for a in authors),
    }


def build_from_xlsx(xlsx_path: str) -> dict:
    """로컬 xlsx 파일을 읽어 저자별 데이터를 집계한다 (폴백)."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl 필요: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    headers = [str(c or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col_map = {}
    for i, h in enumerate(headers):
        hl = h.lower()
        if '상품명' in hl or '제목' in hl:
            col_map['title'] = i
        elif '저자' in hl or '작가' in hl:
            col_map['author'] = i
        elif '출판사' in hl:
            col_map['pub'] = i
        elif '순위' in hl or '순번' in hl:
            col_map['rank'] = i
        elif '수집일' in hl or '날짜' in hl:
            col_map['date'] = i

    if 'author' not in col_map or 'title' not in col_map:
        print(f"저자/상품명 컬럼을 찾을 수 없습니다. 헤더: {headers}", file=sys.stderr)
        sys.exit(1)

    author_map = {}
    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        raw_author = str(row[col_map['author']] or '').strip()
        title = str(row[col_map.get('title', 0)] or '').strip()
        pub = str(row[col_map.get('pub', 0)] or '').strip() if 'pub' in col_map else ''
        rank = 0
        if 'rank' in col_map:
            try:
                rank = int(row[col_map['rank']])
            except (TypeError, ValueError):
                rank = 0
        date_str = str(row[col_map.get('date', 0)] or '').strip() if 'date' in col_map else ''

        parsed = parse_author_field(raw_author)
        for p in parsed:
            if p['role'] in ('역', '감수', '그림', '사진', '편', '원작', '엮음', '해설'):
                continue

            name = p['name']
            if name not in author_map:
                author_map[name] = {'books': {}}
            if title not in author_map[name]['books']:
                author_map[name]['books'][title] = {
                    'title': title, 'pub': pub,
                    'bestRank': rank if rank > 0 else 999,
                    'dates': [], 'appearances': 0
                }
            bk = author_map[name]['books'][title]
            if rank > 0 and rank < bk['bestRank']:
                bk['bestRank'] = rank
            bk['appearances'] += 1
            if date_str and date_str not in bk['dates']:
                bk['dates'].append(date_str)

    wb.close()

    authors = []
    for name, data in author_map.items():
        books = []
        for bk in data['books'].values():
            books.append({
                'title': bk['title'], 'pub': bk['pub'],
                'bestRank': bk['bestRank'], 'days': bk['appearances'],
                'lastDate': sorted(bk['dates'])[-1] if bk['dates'] else ''
            })
        books.sort(key=lambda b: b['bestRank'])
        pubs = sorted(set(b['pub'] for b in books if b['pub']))
        authors.append({
            'name': name, 'books': books, 'pubs': pubs,
            'count': len(books),
            'bestRank': min(b['bestRank'] for b in books) if books else 999,
            'totalDays': sum(b['days'] for b in books),
        })
    authors.sort(key=lambda a: (-a['count'], -a['totalDays'], a['bestRank']))

    return {
        'authors': authors, 'totalRows': row_count,
        'totalAuthors': len(authors),
        'totalBooks': sum(a['count'] for a in authors),
    }


def save_result(result: dict):
    """JSON + JS(브라우저용) 3곳에 저장."""
    os.makedirs(OUT_DIR, exist_ok=True)

    json_path = os.path.join(OUT_DIR, 'authors.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    js_path = os.path.join(OUT_DIR, 'authors.js')
    with open(js_path, 'w', encoding='utf-8') as f:
        f.write('window._AUTHORS_DATA = ')
        json.dump(result, f, ensure_ascii=False)
        f.write(';')

    os.makedirs(os.path.dirname(PANEL_JS_PATH), exist_ok=True)
    with open(PANEL_JS_PATH, 'w', encoding='utf-8') as f:
        f.write('window._AUTHORS_DATA = ')
        json.dump(result, f, ensure_ascii=False)
        f.write(';')

    print(f"✅ 저자 {result['totalAuthors']}명 · 도서 {result['totalBooks']}권")
    print(f"   원본 {result['totalRows']}행 처리")
    print(f"   → {js_path}")
    print(f"   → {PANEL_JS_PATH}")


def main():
    # 1순위: data/yes24/archive.json (GitHub Actions / 자동 수집 데이터)
    if os.path.exists(ARCHIVE_PATH):
        print(f"📖 archive.json 읽는 중...")
        with open(ARCHIVE_PATH, 'r', encoding='utf-8') as f:
            archive = json.load(f)
        if archive.get("snapshots"):
            print(f"   {archive.get('total_days', '?')}일 데이터 ({archive.get('first_date', '?')} ~ {archive.get('last_date', '?')})")
            result = build_from_archive(archive)
            save_result(result)
            return

    # 2순위: 로컬 xlsx 파일 (레거시 폴백)
    xlsx_path = os.path.join(SCRIPT_DIR, '..', '..', '01_yes24_bestseller', 'yes24_IT_베스트셀러_통합.xlsx')
    if os.path.exists(xlsx_path):
        print(f"📖 로컬 xlsx 읽는 중: {os.path.basename(xlsx_path)}...")
        result = build_from_xlsx(xlsx_path)
        save_result(result)
        return

    print("❌ 데이터 소스 없음 (archive.json 또는 로컬 xlsx 필요)", file=sys.stderr)
    sys.exit(1)


if __name__ == '__main__':
    main()
